import csv
import openpyxl
import os
from openpyxl.utils import get_column_letter

def generate_csv_from_txt(txt_file, delimiter):
    with open(txt_file, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    data = []
    for line in lines:
        line = line.strip()
        if delimiter == '-':
            row_data = line.split(delimiter)
        else:
            row_data = [part.strip() for part in line.split(delimiter)]
        data.append(row_data)

    # Utwórz podfolder "kopia_csv" w tym samym katalogu co plik wsadowy
    folder_path = os.path.join(os.path.dirname(txt_file), "kopia_csv")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Zapisz dane do pliku CSV z polskimi znakami w podfolderze "kopia_csv"
    csv_file = os.path.join(folder_path, os.path.basename(txt_file).replace(".txt", ".csv"))
    with open(csv_file, 'w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerows(data)

    print("Dane zostały zapisane do pliku CSV:", csv_file)

    return csv_file

def convert_csv_to_excel(csv_file, excel_file):
    # Otwórz plik Excel przy użyciu openpyxl
    workbook = openpyxl.load_workbook(excel_file)

    # Utwórz nowy arkusz o nazwie takiej jak nazwa pliku wsadowego
    sheet_name = os.path.splitext(os.path.basename(csv_file))[0]
    sheet = workbook.create_sheet(title=sheet_name)

    # Wczytaj dane z pliku CSV
    with open(csv_file, 'r', encoding='utf-8-sig') as file:
        reader = csv.reader(file, delimiter=';')
        for row in reader:
            sheet.append(row)

    # Dopasuj szerokość kolumn na podstawie zawartości
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Zapisz plik Excel
    workbook.save(excel_file)
    print("Dane zostały dodane do pliku Excel.")
